"""
 *packageName    : 
 * fileName       : imageMaker
 * author         : ipeac
 * date           : 2022-08-30
 * description    :
 * ===========================================================
 * DATE              AUTHOR             NOTE
 * -----------------------------------------------------------
 * 2022-08-30        ipeac       최초 생성
 """
import cv2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from matplotlib import colors

img = cv2.imread("SH.jpg", 1)
img = cv2.resize(img, (160, 200), interpolation=cv2.INTER_AREA)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "상지"

h, w, _ = img.shape
for x in range(w):
    sheet.column_dimensions[get_column_letter(x + 1)].width = 5
    for y in range(h):
        b, g, r = img[y, x]
        color_code = colors.to_hex([r / 255, g / 255, b / 255])[1:]
        sheet.cell(row=y + 1, column=x + 1).fill = PatternFill(
            start_color=color_code,
            end_color=color_code,
            fill_type="solid"
        )
        sheet.row_dimensions[y].height = 30

wb.save("./상지사진.xlsx")
